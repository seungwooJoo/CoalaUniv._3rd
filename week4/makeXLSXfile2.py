import openpyxl

wb = openpyxl.Workbook()

sheet1 = wb.active
sheet1.title = "1st sheet"        #시트의 이름을 변경하거나 불러온다. 

sheet2 = wb.creative_sheet('2nd sheet')    #새로운 시트를 원하는 이름으로 생성한다. 

for i in range(1,10):
    sheet1.cell(row=i, column=1).value = i
    sheet2.cell(row=1, column=i).value = i

wb.save('test.xlsx')

#sheet2 = wb["new_sheet"]  >> 원하는 시트를 선택한다. 