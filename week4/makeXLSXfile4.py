'''데이터를 누적해서 저장하기
코드를 실행할 때마다 데이터가 새로 쓰인다
openpyxl.load_workbook('파일이름') >> 소스코드가 있는 폴더에 해당하는 파일이 있으면, 그 파일을 불러온다. 해당하는 파일이 없다면, 에러가 발생'''

import requests
from bs4 import BeautifulSoup
import openpyxl

try:
    wb = openpyxl.load_workbook('naver.xlsx')
    sheet = wb.active
    print('불러오기 성공!')
    
except:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['제목','채널','재생 수','좋아요 수'])
    print('신규 파일 생성!')
    
#...